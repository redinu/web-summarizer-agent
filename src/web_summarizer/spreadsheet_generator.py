"""
Spreadsheet generation for social media content
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
import io

logger = logging.getLogger(__name__)


class SpreadsheetGenerator:
    """Generate spreadsheets from aggregated topic data"""

    def __init__(self):
        pass

    def generate_csv(
        self,
        aggregated_data: Dict,
        output_path: Optional[str] = None,
        include_metadata: bool = True,
    ) -> str:
        """
        Generate CSV spreadsheet from aggregated data

        Args:
            aggregated_data: Data from TopicAggregator
            output_path: Path to save CSV (if None, returns CSV string)
            include_metadata: Include processing metadata columns

        Returns:
            CSV content as string or path to saved file
        """
        if not aggregated_data["data"]:
            raise ValueError("No data to export")

        # Determine columns based on first row and platforms
        first_row = aggregated_data["data"][0]
        platforms = aggregated_data.get("platforms", ["twitter", "linkedin", "facebook"])

        # Base columns
        columns = [
            "source_url",
            "title",
            "category",
            "summary",
            "key_points",
        ]

        # Add platform-specific columns
        for platform in platforms:
            columns.extend([
                f"{platform}_post",
                f"{platform}_hashtags",
                f"{platform}_chars",
            ])

        # Add metadata columns if requested
        if include_metadata:
            columns.extend([
                "word_count",
                "tokens_used",
                "processing_time_ms",
                "timestamp",
            ])

        # Generate CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")

        # Write header
        writer.writeheader()

        # Write data rows
        for row in aggregated_data["data"]:
            writer.writerow(row)

        csv_content = output.getvalue()
        output.close()

        # Save to file if path provided
        if output_path:
            output_path = self._ensure_path(output_path)
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                f.write(csv_content)
            logger.info(f"CSV saved to {output_path}")
            return output_path

        return csv_content

    def generate_excel(
        self,
        aggregated_data: Dict,
        output_path: str,
        include_metadata: bool = True,
    ) -> str:
        """
        Generate Excel spreadsheet with formatting

        Args:
            aggregated_data: Data from TopicAggregator
            output_path: Path to save Excel file
            include_metadata: Include processing metadata

        Returns:
            Path to saved file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        if not aggregated_data["data"]:
            raise ValueError("No data to export")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Social Media Content"

        # Add metadata sheet
        metadata_sheet = wb.create_sheet("Metadata")

        platforms = aggregated_data.get("platforms", ["twitter", "linkedin", "facebook"])

        # Define columns
        columns = [
            ("source_url", "Source URL"),
            ("title", "Title"),
            ("category", "Category"),
            ("summary", "Summary"),
            ("key_points", "Key Points"),
        ]

        # Add platform columns
        for platform in platforms:
            platform_name = platform.capitalize()
            columns.extend([
                (f"{platform}_post", f"{platform_name} Post"),
                (f"{platform}_hashtags", f"{platform_name} Hashtags"),
                (f"{platform}_chars", f"{platform_name} Char Count"),
            ])

        # Add metadata columns
        if include_metadata:
            columns.extend([
                ("word_count", "Word Count"),
                ("tokens_used", "Tokens Used"),
                ("processing_time_ms", "Processing Time (ms)"),
                ("timestamp", "Timestamp"),
            ])

        # Write headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, (field, header) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, row_data in enumerate(aggregated_data["data"], 2):
            for col_idx, (field, _) in enumerate(columns, 1):
                value = row_data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Special formatting for social media posts
                if "_post" in field:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-adjust column widths
        for col_idx, (field, _) in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)

            # Set specific widths for different column types
            if field == "source_url":
                ws.column_dimensions[col_letter].width = 50
            elif field in ["title", "summary"]:
                ws.column_dimensions[col_letter].width = 40
            elif field == "key_points":
                ws.column_dimensions[col_letter].width = 60
            elif "_post" in field:
                ws.column_dimensions[col_letter].width = 70
            elif "_hashtags" in field:
                ws.column_dimensions[col_letter].width = 30
            elif "_chars" in field:
                ws.column_dimensions[col_letter].width = 12
            else:
                ws.column_dimensions[col_letter].width = 15

        # Add freeze panes (freeze first row)
        ws.freeze_panes = "A2"

        # Add metadata to metadata sheet
        metadata_sheet["A1"] = "Topic"
        metadata_sheet["B1"] = aggregated_data["topic"]
        metadata_sheet["A2"] = "Total Sources"
        metadata_sheet["B2"] = aggregated_data["total_sources"]
        metadata_sheet["A3"] = "Successful Summaries"
        metadata_sheet["B3"] = aggregated_data["successful_summaries"]
        metadata_sheet["A4"] = "Failed Summaries"
        metadata_sheet["B4"] = aggregated_data["failed_summaries"]
        metadata_sheet["A5"] = "Generated At"
        metadata_sheet["B5"] = aggregated_data["generated_at"]
        metadata_sheet["A6"] = "Platforms"
        metadata_sheet["B6"] = ", ".join(platforms)

        # Style metadata sheet
        for row in range(1, 7):
            metadata_sheet[f"A{row}"].font = Font(bold=True)
            metadata_sheet.column_dimensions["A"].width = 20
            metadata_sheet.column_dimensions["B"].width = 40

        # Save workbook
        output_path = self._ensure_path(output_path)
        wb.save(output_path)
        logger.info(f"Excel file saved to {output_path}")

        return output_path

    def _ensure_path(self, path: str) -> str:
        """Ensure output directory exists and return full path"""
        path_obj = Path(path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        return str(path_obj)

    def generate_pdf(
        self,
        aggregated_data: Dict,
        output_path: str,
        include_metadata: bool = True,
    ) -> str:
        """
        Generate PDF report with formatting

        Args:
            aggregated_data: Data from TopicAggregator
            output_path: Path to save PDF file
            include_metadata: Include processing metadata

        Returns:
            Path to saved file
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
        except ImportError:
            raise ImportError(
                "reportlab is required for PDF export. Install with: pip install reportlab"
            )

        if not aggregated_data["data"]:
            raise ValueError("No data to export")

        output_path = self._ensure_path(output_path)

        # Create PDF document
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=1*inch,
            bottomMargin=0.75*inch,
        )

        # Build story (content)
        story = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1a73e8'),
            spaceAfter=12,
            spaceBefore=12,
        )

        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            spaceAfter=10,
        )

        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['BodyText'],
            fontSize=11,
            leading=14,
            alignment=TA_JUSTIFY,
        )

        # Title
        story.append(Paragraph(f"Research Report: {aggregated_data['topic']}", title_style))
        story.append(Spacer(1, 0.3*inch))

        # Metadata
        story.append(Paragraph(f"<b>Generated:</b> {aggregated_data['generated_at']}", body_style))
        story.append(Paragraph(f"<b>Total Sources:</b> {aggregated_data['total_sources']}", body_style))
        story.append(Paragraph(f"<b>Successful Summaries:</b> {aggregated_data['successful_summaries']}", body_style))
        story.append(Paragraph(f"<b>Platforms:</b> {', '.join(aggregated_data.get('platforms', []))}", body_style))
        story.append(Spacer(1, 0.3*inch))

        # Master Summary
        if aggregated_data.get('master_summary'):
            story.append(Paragraph("Master Summary", heading_style))
            story.append(Paragraph(aggregated_data['master_summary'], body_style))
            story.append(Spacer(1, 0.2*inch))

        # Social Media Posts
        if aggregated_data.get('social_media_posts'):
            story.append(Paragraph("Social Media Posts (Ready to Publish)", heading_style))
            story.append(Spacer(1, 0.1*inch))

            for platform, post_content in aggregated_data['social_media_posts'].items():
                story.append(Paragraph(f"<b>{platform.upper()}</b>", subheading_style))
                # Replace newlines with <br/> tags for PDF
                formatted_post = post_content.replace('\n', '<br/>')
                story.append(Paragraph(formatted_post, body_style))
                story.append(Paragraph(f"<i>{len(post_content)} characters</i>", body_style))
                story.append(Spacer(1, 0.15*inch))

        # Source Articles
        story.append(PageBreak())
        story.append(Paragraph("Source Articles & Summaries", heading_style))
        story.append(Spacer(1, 0.1*inch))

        for idx, row in enumerate(aggregated_data['data'], 1):
            story.append(Paragraph(f"<b>Article {idx}: {row['title']}</b>", subheading_style))
            story.append(Paragraph(f"<i>Source:</i> <a href='{row['source_url']}'>{row['source_url']}</a>", body_style))
            story.append(Spacer(1, 0.05*inch))
            story.append(Paragraph(f"<b>Summary:</b>", body_style))
            story.append(Paragraph(row['summary'], body_style))
            story.append(Spacer(1, 0.05*inch))
            story.append(Paragraph(f"<b>Key Points:</b> {row['key_points']}", body_style))
            story.append(Spacer(1, 0.2*inch))

            # Add page break every 2 articles to avoid cramming
            if idx % 2 == 0 and idx < len(aggregated_data['data']):
                story.append(PageBreak())

        # Build PDF
        doc.build(story)
        logger.info(f"PDF saved to {output_path}")

        return output_path

    def preview_csv(self, aggregated_data: Dict, max_rows: int = 5) -> str:
        """
        Generate a preview of the CSV data

        Args:
            aggregated_data: Data from TopicAggregator
            max_rows: Maximum number of rows to preview

        Returns:
            Preview string
        """
        if not aggregated_data["data"]:
            return "No data available"

        preview_data = {
            **aggregated_data,
            "data": aggregated_data["data"][:max_rows],
        }

        csv_content = self.generate_csv(preview_data, output_path=None)

        preview = f"Preview (showing {min(max_rows, len(aggregated_data['data']))} of {len(aggregated_data['data'])} rows):\n\n"
        preview += csv_content

        return preview
